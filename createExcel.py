from fastapi import FastAPI, HTTPException, BackgroundTasks
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import os
from typing import Optional
from datetime import datetime
import logging
import pytz

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

# MySQL connection
DATABASE_URL = "mysql://root:123456@localhost:3306/calculatorDev"
engine = create_engine(DATABASE_URL, pool_size=5, max_overflow=10)
# Get the directory where the script is located
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(SCRIPT_DIR, "reports")
malaysia_tz = pytz.timezone('Asia/Kuala_Lumpur')

def get_accounts():
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT id, name, accountHolder, accountNo 
                FROM Bank
                ORDER BY name
            """))
            accounts = result.fetchall()
            logger.info(f"Retrieved {len(accounts)} accounts")
            return accounts
    except Exception as e:
        logger.error(f"Error retrieving accounts: {str(e)}")
        raise

def get_banks():
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
               SELECT 
                    IFNULL(SUM(a.amount), 0) AS total,  -- Replace NULL with 0
                    b.name,
                    b.accountNo
                FROM 
                    task AS a 
                RIGHT JOIN 
                    bank AS b 
                ON 
                    a.bankId = b.id 
                AND 
                    a.createdAt BETWEEN '2025-02-11 00:00:00' AND '2025-02-11 23:59:59'
                GROUP BY 
                    b.id;
            """))
            accounts = result.fetchall()
            logger.info(f"Retrieved {len(accounts)} accounts")
            return accounts
    except Exception as e:
        logger.error(f"Error retrieving accounts: {str(e)}")
        raise

def get_transactions(bank_id):
    try:
        with engine.connect() as conn:
            BATCH_SIZE = 1000
            all_transactions = []
            offset = 0
            
            while True:
                result = conn.execute(text("""
                    SELECT amount, type, status, createdAt,reason
                    FROM Task
                    WHERE bankId = :bank_id
                    ORDER BY createdAt DESC
                    LIMIT :limit OFFSET :offset
                """), {
                    "bank_id": bank_id,
                    "limit": BATCH_SIZE,
                    "offset": offset
                })
                
                batch = result.fetchall()
                if not batch:
                    break
                    
                all_transactions.extend(batch)
                offset += BATCH_SIZE
                
                if len(batch) < BATCH_SIZE:
                    break
            
            logger.info(f"Retrieved {len(all_transactions)} transactions for bank_id {bank_id}")
            return all_transactions
    except Exception as e:
        logger.error(f"Error retrieving transactions for bank_id {bank_id}: {str(e)}")
        raise
def get_branchs():
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
               SELECT 
                    IFNULL(SUM(a.amount), 0) AS total,  -- Replace NULL with 0
                    b.code 
                FROM 
                    task AS a 
                RIGHT JOIN 
                    branch AS b 
                ON 
                    a.branchId = b.id 
                AND a.createdAt BETWEEN '2025-02-11 00:00:00' AND '2025-02-11 23:59:59'
                GROUP BY 
                b.id;
            """))
            accounts = result.fetchall()
            logger.info(f"Retrieved {len(accounts)} branchs")
            return accounts
    except Exception as e:
        logger.error(f"Error retrieving branchs: {str(e)}")
        raise

def create_excel_report(report_id: str):
    try:
        logger.info(f"Starting report generation for report_id: {report_id}")
        wb = Workbook()
        ws = wb.active
        
        # Styles
        light_blue = "B6D7E8"
        light_pink = "FFD9D9"
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get accounts
        accounts = get_accounts()
        if not accounts:
            raise ValueError("No accounts found")
        
        # Create directory with full permissions
        report_dir =os.path.join(SCRIPT_DIR, "reports")
        os.makedirs(report_dir, mode=0o777, exist_ok=True)
        
        # Create top row with zeros
        for col in range(1, len(accounts) * 6 + 3):
            cell = ws.cell(row=1, column=col, value=0)
            cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
            cell.border = thin_border
        
        # Create headers with account information
        ws.cell(row=2, column=1, value="进出账记录").fill = PatternFill(start_color=light_pink, end_color=light_pink, fill_type="solid")
        ws.cell(row=2, column=2, ).fill = PatternFill(start_color=light_pink, end_color=light_pink, fill_type="solid")
        cell = ws.cell(row=2, column=1)
        cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=2,start_column=1,end_row=2, end_column=2)

        current_col = 3
        for account in accounts:
            cell = ws.cell(row=2, column=current_col)
            # Handle potential None values
            name = getattr(account, 'name', '') or ''
            account_no = getattr(account, 'accountNo', '') or ''
            cell.value = f"{name} {account_no}".strip()
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
            cell.border = thin_border
            ws.merge_cells(start_row=2, start_column=current_col, end_row=2, end_column=current_col + 4)
            current_col += 6
        
        ws.cell(row=3, column=1, value="底线").fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
        # Add column headers
        headers = ["Date", "Note", "Withdraw", "deposit", "Time"]
        current_col = 3
        for _ in accounts:
            for header in headers:
                cell = ws.cell(row=3, column=current_col, value=header)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
                current_col += 1
            current_col += 1
        
        # Process transactions
        current_col = 3
        #   ROW FOR BANK LIST
        row_bank = 4
        for account in accounts:
            try:
                # Add START
                cell = ws.cell(row=4, column=current_col, value="START")
                cell.border = thin_border
               
                # Get and process transactions for this account
                transactions = get_transactions(account.id)
                total_rowFormula =0
                for idx, trans in enumerate(transactions, start=4):
                    try:
                        # Handle potential None or invalid datetime values
                        created_at = getattr(trans, 'createdAt', None)
                        typeTask =  getattr(trans, 'type', '')
                        if created_at:
                            if created_at.tzinfo is None:
                                created_at = pytz.utc.localize(created_at)

                            malaysia_time = created_at.astimezone(malaysia_tz)    
                            date_str = malaysia_time.strftime('%Y-%m-%d')
                            time_str = malaysia_time.strftime('%H:%M:%S')
                        else:
                            date_str = ''
                            time_str = ''
                            
                        ws.cell(row=idx, column=current_col).value = date_str
                        ws.cell(row=idx, column=current_col + 1).value = getattr(trans, 'reason', '')
                        if typeTask == "WITHDRAW" :
                             ws.cell(row=idx, column=current_col + 2).value = getattr(trans, 'amount', 0)
                        else:
                            ws.cell(row=idx, column=current_col + 3).value = getattr(trans, 'amount', '')
                        ws.cell(row=idx, column=current_col + 4).value = time_str
                        total_rowFormula = idx
                    except Exception as e:
                        logger.error(f"Error processing transaction: {str(e)}")
                        continue
                #SUM FOR WITHDRAW
                cell_withdraw = get_column_letter(current_col + 2)
                cell_withdraw_formula = f"=SUM({cell_withdraw}4:{cell_withdraw}{total_rowFormula})"
                cell = ws.cell(row=1, column=current_col+2, value=cell_withdraw_formula)
                #SUM FOR DEPOSIT
                cell_deposit = get_column_letter(current_col + 3)
                cell_DEPOSIT_formula = f"=SUM({cell_deposit}4:{cell_deposit}{total_rowFormula})"
                cell = ws.cell(row=1, column=current_col+3, value=cell_DEPOSIT_formula)
                #LAST ROW FOR EACH BANK
                cell = ws.cell(row=total_rowFormula +1 , column=current_col, value="TOTAL")
                cell = ws.cell(row=total_rowFormula +1 , column=current_col+2, value=cell_withdraw_formula)
                cell = ws.cell(row=total_rowFormula +1 , column=current_col+3, value=cell_DEPOSIT_formula)
                #CENTER TOTAL
                cell = ws.cell(row=total_rowFormula +1, column=current_col)
                cell.alignment = Alignment(horizontal='center')
                ws.merge_cells(start_row=total_rowFormula +1,start_column=current_col,end_row=total_rowFormula +1, end_column=current_col+1)
                #TOTAL ALL TASK
                formula_Total_All = f"=({cell_withdraw}{total_rowFormula+1} - {cell_deposit}{total_rowFormula+1}) * -1"
                cell = ws.cell(row=1 , column=current_col+1, value=formula_Total_All)
                # ADD BANK
                name_Bank = get_column_letter(current_col )
                value_total_Bank = get_column_letter(current_col +1)
                ws.cell(row=row_bank,column= 1 ,value=f"={name_Bank}2")
                ws.cell(row=row_bank,column= 2 ,value=f"={value_total_Bank}1")
                row_bank += 1
                logger.info(f"row_bank: {value_total_Bank} ")
            except Exception as e:
                logger.error(f"Error processing account {getattr(account, 'id', 'unknown')}: {str(e)}")
                continue
                
            current_col += 6
        
        array_length = len(accounts)
        last_row_Bank = array_length
       
        banks = get_banks()
        # Add left side account list
        # for idx, bank in enumerate(banks, start=4):
        #     try:
        #         # Handle potential None values
        #         name = getattr(bank, 'name', '') or ''
        #         account_no = getattr(bank, 'accountNo', '') or ''
        #         total = getattr(bank, 'total', '') or ''
                
        #         # Account name with account holder
        #         cell = ws.cell(row=idx, column=1, value=f"{name} ({account_no})".strip())
        #         cell.fill = PatternFill(start_color=light_pink, end_color=light_pink, fill_type="solid")
        #         cell.border = thin_border
        #         cell.alignment = Alignment(wrap_text=True)
        #         # Account number
        #         cell = ws.cell(row=idx, column=2, value=total)
        #         cell.fill = PatternFill(start_color=light_pink, end_color=light_pink, fill_type="solid")
        #         cell.border = thin_border
        #         last_row_Bank = idx
        #     except Exception as e:
        #         logger.error(f"Error adding account to left side list: {str(e)}")
        #         continue
        cell = ws.cell(row=last_row_Bank+ last_row_Bank+1, column=1, value="cash")
        cell = ws.cell(row=last_row_Bank+ last_row_Bank+1, column=2, value=f"=SUM(B4:B{3+last_row_Bank})")
        cell = ws.cell(row=last_row_Bank+ last_row_Bank+2, column=1, value="自动计算")
        cell = ws.cell(row=last_row_Bank+ last_row_Bank+3, column=1, value="今日未领")
        cell = ws.cell(row=last_row_Bank+ last_row_Bank+4, column=1, value="TOTAL")
        cell = ws.cell(row=last_row_Bank+ last_row_Bank+5, column=1, value="场口已领")
        cell = ws.cell(row=last_row_Bank+ last_row_Bank+6, column=1, value="")
        cell = ws.cell(row=last_row_Bank+ last_row_Bank+7, column=1, value="自动计算")

        branchs = get_branchs()
        if not branchs:
            raise ValueError("No branchs found")

        for idx, branch in enumerate(branchs, start=last_row_Bank+ last_row_Bank+8):
            try:
                # Handle potential None values
                code = getattr(branch, 'code', '') or ''
               
                total = getattr(branch, 'total', '') or ''
                
                # Account name with account holder
                cell = ws.cell(row=idx, column=1, value=f"{code}".strip())
                cell.fill = PatternFill(start_color=light_pink, end_color=light_pink, fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)
                # Account number
                cell = ws.cell(row=idx, column=2, value=total)
                cell.fill = PatternFill(start_color=light_pink, end_color=light_pink, fill_type="solid")
                cell.border = thin_border
                last_row_Bank = idx
            except Exception as e:
                logger.error(f"Error adding account to left side list: {str(e)}")
                continue

        # Set column widths
        for col in range(1, current_col):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save file with full path
        filename = f"financial_report_{report_id}.xlsx"
        filepath = os.path.join(report_dir, filename)
        
        logger.info(f"Attempting to save report to: {filepath}")
        wb.save(filepath)
        
        # Verify file was created
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Failed to create file at {filepath}")
            
        # Set file permissions
        os.chmod(filepath, 0o666)
        
        logger.info(f"Successfully created report: {filepath}")
        return filename
        
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        raise

@app.post("/generate-excel")
async def generate_excel(background_tasks: BackgroundTasks):
    try:
        report_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        logger.info(f"Starting background task for report_id: {report_id}")
        background_tasks.add_task(create_excel_report, report_id)
        
        return {
            "status": "processing",
            "report_id": report_id,
            "message": "Report generation started"
        }
    except Exception as e:
        logger.error(f"Error in generate-excel endpoint: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/report-status/{report_id}")
async def check_status(report_id: str):
    try:
        filepath = os.path.join("/tmp/reports", f"financial_report_{report_id}.xlsx")
        
        if os.path.exists(filepath):
            return {
                "status": "completed",
                "filename": f"financial_report_{report_id}.xlsx"
            }
        return {
            "status": "processing"
        }
    except Exception as e:
        logger.error(f"Error checking report status: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))