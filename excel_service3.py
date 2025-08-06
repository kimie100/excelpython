import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font,Color
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
import pytz
from datetime import datetime

from config import logger, REPORTS_DIR, malaysia_tz,color_code_mapping
from data_service import get_accounts, get_banks, get_transactions,get_transactions2, get_branches,get_total

def create_excel_report3(report_id: str, date_range=None, reportdata: dict = None):
    try:
        if not date_range:
            today = datetime.now().strftime('%Y-%m-%d')
            date_range = (f"{today} 00:00:00", f"{today} 23:59:59")

        logger.info(f"Starting report generation for report_id: {report_id}")
        wb = Workbook()
        ws = wb.active

        # Styles
        light_blue = "B6D7E8"
        light_pink = "FFD9D9"
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Extract data from reportdata
        accounts = reportdata.get("result", []) if reportdata else []
        branch_data = reportdata.get("branch", []) if reportdata else []
        totals = reportdata.get("total", {}) if reportdata else {
            "pendingTotal": 0,
            "completeTotal": 0,
            "grandTotal": 0
        }

        if not accounts:
            raise ValueError("No account data found in reportdata")

        # Create top row with zeros
        for col in range(1, len(accounts) * 7 + 3):
            cell = ws.cell(row=1, column=col, value=0)
            cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
            cell.border = thin_border

        # Header row for account names
        ws.cell(row=2, column=1, value="进出账记录").fill = PatternFill(
            start_color=light_pink, end_color=light_pink, fill_type="solid")
        ws.cell(row=2, column=2).fill = PatternFill(
            start_color=light_pink, end_color=light_pink, fill_type="solid")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

        current_col = 3
        for account in accounts:
            name = account.get("name", "") or ""
            account_no = account.get("accountNo", "") or ""
            cell = ws.cell(row=2, column=current_col)

            # Rich text: black name + red account number
            red = InlineFont(color='FF0000')
            black = InlineFont(color='000000')
            rich_string = CellRichText([
                TextBlock(black, name + ' '),
                TextBlock(red, account_no)
            ])
            cell.value = rich_string
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            ws.merge_cells(start_row=2, start_column=current_col, end_row=2, end_column=current_col + 5)
            current_col += 7

        # Sub-headers: Date, Note, Name, Withdraw, Deposit, Time
        headers = ["Date", "Note", "Name", "Withdraw", "Deposit", "Time"]
        current_col = 3
        for account in accounts:
            for i, header in enumerate(headers):
                cell = ws.cell(row=3, column=current_col + i, value=header)
                cell.font = Font(bold=True)
                cell.border = thin_border
            current_col += 7

        # Process transactions per account
        current_col = 3
        max_row = 4  # Track max row used across all banks

        for account in accounts:
            tasks = account.get("tasks", [])
            total_rowFormula = 3  # Will be updated per account

            row_idx = 4
            for task in tasks:
                try:
                    created_at = task.get("date")
                    if created_at:
                        # Parse ISO format
                        dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                        dt = dt.astimezone(malaysia_tz)
                        date_str = dt.strftime('%d-%m-%Y')
                        time_str = dt.strftime('%H:%M:%S')
                    else:
                        date_str = time_str = ''

                    code = task.get("branchs", {}).get("code", "")
                    color_Branch = color_code_mapping.get(code, "FFFFFF")

                    # Write data
                    ws.cell(row=row_idx, column=current_col, value=date_str).border = thin_border
                    ws.cell(row=row_idx, column=current_col + 1, value=code).border = thin_border
                    ws.cell(row=row_idx, column=current_col + 2, value=task.get("name", "")).alignment = Alignment(wrap_text=True)
                    ws.cell(row=row_idx, column=current_col + 2).border = thin_border

                    if task.get("type") == "WITHDRAW":
                        ws.cell(row=row_idx, column=current_col + 3, value=task.get("amount", 0)).border = thin_border
                    else:
                        ws.cell(row=row_idx, column=current_col + 4, value=task.get("amount", 0)).border = thin_border

                    ws.cell(row=row_idx, column=current_col + 5, value=time_str).border = thin_border

                    # Fill color by branch
                    for i in range(6):
                        ws.cell(row=row_idx, column=current_col + i).fill = PatternFill(
                            start_color=color_Branch, end_color=color_Branch, fill_type="solid")

                    row_idx += 1
                    total_rowFormula = row_idx - 1
                except Exception as e:
                    logger.error(f"Error processing task {task.get('id')}: {str(e)}")
                    continue

            # Add SUM formulas for this bank
            if total_rowFormula >= 4:
                withdraw_col = get_column_letter(current_col + 3)
                deposit_col = get_column_letter(current_col + 4)

                withdraw_sum = f"=SUM({withdraw_col}4:{withdraw_col}{total_rowFormula})"
                deposit_sum = f"=SUM({deposit_col}4:{deposit_col}{total_rowFormula})"

                ws.cell(row=1, column=current_col + 3, value=withdraw_sum)
                ws.cell(row=1, column=current_col + 4, value=deposit_sum)

                # Total row
                ws.cell(row=total_rowFormula + 1, column=current_col, value="TOTAL").alignment = Alignment(horizontal='center')
                ws.merge_cells(start_row=total_rowFormula + 1, start_column=current_col, end_row=total_rowFormula + 1, end_column=current_col + 1)
                ws.cell(row=total_rowFormula + 1, column=current_col + 3, value=withdraw_sum)
                ws.cell(row=total_rowFormula + 1, column=current_col + 4, value=deposit_sum)

                # Update max_row
                max_row = max(max_row, total_rowFormula + 1)

            current_col += 7

        # === Summary Section ===
        summary_start_row = max_row + 3

        # Cash
        ws.cell(row=summary_start_row, column=1, value="cash").alignment = Alignment(horizontal='center')
        ws.cell(row=summary_start_row, column=1).fill = PatternFill(start_color="DCE6F1", fill_type="solid")
        ws.cell(row=summary_start_row, column=2).fill = PatternFill(start_color="FFFF66", fill_type="solid")
        ws.cell(row=summary_start_row, column=2, value=f"=SUM(B4:B{summary_start_row - 1})").border = thin_border

        # 月 (merge)
        ws.merge_cells(start_row=summary_start_row + 1, start_column=1, end_row=summary_start_row + 1, end_column=2)
        ws.cell(row=summary_start_row + 1, column=1, value="月").alignment = Alignment(horizontal='center')

        # Auto calculation
        ws.cell(row=summary_start_row + 2, column=1, value="自动计算").fill = PatternFill(start_color="C5D9F1", fill_type="solid")
        ws.cell(row=summary_start_row + 2, column=2, value=totals.get("completeTotal", 0)).fill = PatternFill(start_color="DCE6F1", fill_type="solid")

        ws.cell(row=summary_start_row + 3, column=1, value="今日未领").fill = PatternFill(start_color="C5D9F1", fill_type="solid")
        ws.cell(row=summary_start_row + 3, column=2, value=totals.get("pendingTotal", 0)).fill = PatternFill(start_color="DCE6F1", fill_type="solid")

        ws.cell(row=summary_start_row + 4, column=1, value="TOTAL").fill = PatternFill(start_color="C5D9F1", fill_type="solid")
        ws.cell(row=summary_start_row + 4, column=2, value=f"=SUM(B{summary_start_row + 2}:B{summary_start_row + 3})").fill = PatternFill(start_color="DCE6F1", fill_type="solid")

        ws.cell(row=summary_start_row + 5, column=1, value="场口已领").fill = PatternFill(start_color="FFFFCC", fill_type="solid")
        ws.cell(row=summary_start_row + 5, column=2).fill = PatternFill(start_color="FFFFCC", fill_type="solid")

        ws.cell(row=summary_start_row + 7, column=1, value="自动计算").border = thin_border
        ws.cell(row=summary_start_row + 7, column=2).border = thin_border

        # Apply borders
        for r in [summary_start_row, summary_start_row + 2, summary_start_row + 3, summary_start_row + 4, summary_start_row + 5, summary_start_row + 7]:
            ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=2).border = thin_border

        # === Branch Data ===
        branch_start_row = summary_start_row + 10
        for idx, branch in enumerate(branch_data):
            code = branch.get("code", "")
            amount = branch.get("amount", 0)
            color_branch = color_code_mapping.get(code, "FFFFFF")

            row = branch_start_row + idx
            ws.cell(row=row, column=1, value=code).fill = PatternFill(start_color=color_branch, fill_type="solid")
            ws.cell(row=row, column=2, value=amount).fill = PatternFill(start_color=color_branch, fill_type="solid")
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2).border = thin_border

        # Final SUM for branch
        last_branch_row = branch_start_row + len(branch_data) - 1
        ws.cell(row=branch_start_row - 1, column=2, value=f"=SUM(B{branch_start_row}:B{last_branch_row})")

        # Set column widths
        for col in range(1, current_col):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Save file
        filename = f"financial_report_{report_id}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)
        wb.save(filepath)

        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Failed to create file at {filepath}")
        os.chmod(filepath, 0o666)

        logger.info(f"Successfully created report: {filepath}")
        return filename

    except Exception as e:
        logger.error(f"Error generating report: {str(e)}", exc_info=True)
        raise