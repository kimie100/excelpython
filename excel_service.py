import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font,Color
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
import pytz
from datetime import datetime

from config import logger, REPORTS_DIR, malaysia_tz,color_code_mapping
from data_service import get_accounts, get_banks, get_transactions,get_transactions2, get_branches,get_total

def create_excel_report(report_id: str,date_range=None):
    try:
        if not date_range:
            today = datetime.now().strftime('%Y-%m-%d')
            date_range = (f"{today} 00:00:00", f"{today} 23:59:59")
        logger.info(f"Starting report generation for report_id: {report_id}")
        wb = Workbook()
        ws = wb.active
        
        # Styles
        light_blue = "B6D7E8"
        light_pink = "FFD9D9"
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get accounts
        accounts = get_accounts()
        if not accounts:
            raise ValueError("No accounts found")
        
        # Create top row with zeros
        for col in range(1, len(accounts) * 6 + 3):
            cell = ws.cell(row=1, column=col, value=0)
            cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
            cell.border = thin_border
        
        # Create headers with account information
        ws.cell(row=2, column=1, value="进出账记录").fill = PatternFill(start_color=light_pink, end_color=light_pink, fill_type="solid")
        ws.cell(row=2, column=2, ).fill = PatternFill(start_color=light_pink, end_color=light_pink, fill_type="solid")
        cell = ws.cell(row=2, column=1)
        cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=2,start_column=1,end_row=2, end_column=2)

        current_col = 3
        for account in accounts:
            cell = ws.cell(row=2, column=current_col)
            # Handle potential None values
            name = getattr(account, 'name', '') or ''
            account_no = getattr(account, 'accountNo', '') or ''
            if isinstance(account_no, int):
                account_no = str(account_no)
            red = InlineFont(color='FF0000')
            black  = InlineFont(color='000000')
            rich_string1 = CellRichText([TextBlock(black, name + ' '),  TextBlock(red, account_no)])
            # cell.value = f"{name} {account_no}".strip()
            cell.value = rich_string1
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            # cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
            cell.border = thin_border
            ws.merge_cells(start_row=2, start_column=current_col, end_row=2, end_column=current_col + 4)
            current_col += 6
        
        ws.cell(row=3, column=1, value="底线").fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
        # Add column headers
        headers = ["Date", "Note", "Withdraw", "deposit", "Time"]
        current_col = 3
        for _ in accounts:
            for header in headers:
                cell = ws.cell(row=3, column=current_col, value=header)
                cell.font = Font(bold=True)
                cell.border = thin_border
                # cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
                current_col += 1
            current_col += 1
        
        # Process transactions
        current_col = 3
        #   ROW FOR BANK LIST
        row_bank = 4
        for account in accounts:
            try:
                # Add START
                # cell = ws.cell(row=4, column=current_col, value="START")
                cell.border = thin_border
               
                # Get and process transactions for this account
                # transactions = get_transactions(account.id,date_range)
                transactions = get_transactions2(account.id,date_range)
                 # ADD BANK 
                color_bank = "FFC080"
                color_value_bank ="FFC0C0"
                name_Bank = get_column_letter(current_col)
                value_total_Bank = get_column_letter(current_col + 1)
                ws.cell(row=row_bank, column=1, value=f"={name_Bank}2")
                ws.cell(row=row_bank, column=1,).fill = PatternFill(start_color=color_bank, end_color=color_bank, fill_type="solid")
                ws.cell(row=row_bank, column=1,).border  = thin_border
                ws.cell(row=row_bank, column=2, value=f"={value_total_Bank}1")
                ws.cell(row=row_bank, column=2,).fill = PatternFill(start_color=color_value_bank, end_color=color_value_bank, fill_type="solid")
                ws.cell(row=row_bank, column=2,).border  = thin_border
                row_bank += 1
                # logger.info(f"row bank name{name_Bank}: {value_total_Bank}")
                # logger.info(f"row {row_bank}: {row_bank}")
                if not transactions:
                    logger.info(f"list tast for bank ${account.id}: {transactions}")
                else:
                    total_rowFormula = 0
                    for idx, trans in enumerate(transactions, start=4):
                        try:
                            # Handle potential None or invalid datetime values
                            
                            created_at = getattr(trans, 'updatedAt', None)
                            logger.info(f"updatedAt: {created_at}")
                            typeTask = getattr(trans, 'type', '')
                            logger.info(f"row {typeTask}: {typeTask}")
                            code = getattr(trans, 'code', '')
                            if code is not None:
                                logger.info(f"ada color: {code}")
                                color_Branch = get_color_by_code(code)
                            else:
                                color_Branch = 'FFFFFF'
                            if created_at:
                                if created_at.tzinfo is None:
                                    created_at = pytz.utc.localize(created_at)

                                malaysia_time = created_at.astimezone(malaysia_tz)    
                                date_str = malaysia_time.strftime('%Y-%m-%d')
                                time_str = malaysia_time.strftime('%H:%M:%S')
                            else:
                                date_str = ''
                                time_str = ''
                                
                            ws.cell(row=idx, column=current_col).value = date_str
                            ws.cell(row=idx, column=current_col + 1).value = getattr(trans, 'status', '')

                            
                            ws.cell(row=idx, column=current_col + 1).border = thin_border
                            if typeTask == "WITHDRAW":
                                ws.cell(row=idx, column=current_col + 2).value = getattr(trans, 'amount', 0)
                            else:
                                ws.cell(row=idx, column=current_col + 3).value = getattr(trans, 'amount', '')
                            
                            ws.cell(row=idx, column=current_col + 2).border = thin_border
                            ws.cell(row=idx, column=current_col + 3).border = thin_border
                            ws.cell(row=idx, column=current_col + 4).value = time_str
                            #color row
                            ws.cell(row=idx, column=current_col).fill = PatternFill(start_color=color_Branch, end_color=color_Branch, fill_type="solid")
                            ws.cell(row=idx, column=current_col + 1).fill = PatternFill(start_color=color_Branch, end_color=color_Branch, fill_type="solid")
                            ws.cell(row=idx, column=current_col + 2).fill = PatternFill(start_color=color_Branch, end_color=color_Branch, fill_type="solid")
                            ws.cell(row=idx, column=current_col + 3).fill = PatternFill(start_color=color_Branch, end_color=color_Branch, fill_type="solid")
                            ws.cell(row=idx, column=current_col + 4).fill = PatternFill(start_color=color_Branch, end_color=color_Branch, fill_type="solid")
                            total_rowFormula = idx
                        except Exception as e:
                            logger.error(f"Error processing transaction: {str(e)}")
                            continue
                            
                    # Add formulas and totals
                    _add_bank_summary(ws, current_col, total_rowFormula)
                    
                   
                    logger.info(f"row_bank: {value_total_Bank}")
            except Exception as e:
                logger.error(f"Error processing account {getattr(account, 'id', 'unknown')}: {str(e)}")
                continue
                
            current_col += 6
        
        array_length = len(accounts)
        last_row_Bank = array_length
        logger.info(f"last_row_Bank: {last_row_Bank}")
        # Add bank summary section
        _add_bank_list_summary(ws, last_row_Bank,date_range)
        
        # Add branch data
        _add_branch_data(ws, last_row_Bank,date_range)
        
        # Set column widths
        for col in range(1, current_col):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save file with full path
        filename = f"financial_report_{report_id}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)
        
        logger.info(f"Attempting to save report to: {filepath}")
        wb.save(filepath)
        
        # Verify file was created and set permissions
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Failed to create file at {filepath}")
            
        os.chmod(filepath, 0o666)
        
        logger.info(f"Successfully created report: {filepath}")
        return filename
        
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        raise

def _add_bank_summary(ws, current_col, total_rowFormula):
    """Helper to add formulas and totals for each bank column"""
    #SUM FOR WITHDRAW
    cell_withdraw = get_column_letter(current_col + 2)
    cell_withdraw_formula = f"=SUM({cell_withdraw}4:{cell_withdraw}{total_rowFormula})"
    ws.cell(row=1, column=current_col+2, value=cell_withdraw_formula)
    
    #SUM FOR DEPOSIT
    cell_deposit = get_column_letter(current_col + 3)
    cell_DEPOSIT_formula = f"=SUM({cell_deposit}4:{cell_deposit}{total_rowFormula})"
    ws.cell(row=1, column=current_col+3, value=cell_DEPOSIT_formula)
    
    #LAST ROW FOR EACH BANK
    ws.cell(row=total_rowFormula + 1, column=current_col, value="TOTAL")
    ws.cell(row=total_rowFormula + 1, column=current_col+2, value=cell_withdraw_formula)
    ws.cell(row=total_rowFormula + 1, column=current_col+3, value=cell_DEPOSIT_formula)
    
    #CENTER TOTAL
    cell = ws.cell(row=total_rowFormula + 1, column=current_col)
    cell.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=total_rowFormula + 1, start_column=current_col, 
                   end_row=total_rowFormula + 1, end_column=current_col+1)
                   
    #TOTAL ALL TASK
    formula_Total_All = f"=({cell_withdraw}{total_rowFormula+1} - {cell_deposit}{total_rowFormula+1}) * -1"
    ws.cell(row=1, column=current_col+1, value=formula_Total_All)

def _add_bank_list_summary(ws, last_row_Bank,date_range):
    """Add summary sections for banks"""
    light_pink = "FFD9D9"
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    summary_row = 4 + last_row_Bank 
    totals = get_total(date_range)
    ws.cell(row=summary_row, column=1, value="cash")
    #center cash C0C0FF
    ws.cell(row=summary_row, column=1,).alignment = Alignment(horizontal='center')
    ws.cell(row=summary_row, column=1,).fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    ws.cell(row=summary_row, column=2, value=f"=SUM(B4:B{3+last_row_Bank})")
    ws.cell(row=summary_row, column=2,).fill = PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type="solid")
    ws.cell(row=summary_row +1, column=1, value="月")
    #merge 月
    ws.merge_cells(start_row=summary_row +1, start_column=1, end_row=summary_row +1, end_column=2)
    # center 月
    ws.cell(row=summary_row +1, column=1,).alignment = Alignment(horizontal='center')

    ws.cell(row=summary_row + 2, column=1, value="自动计算")
    ws.cell(row=summary_row + 2, column=2, value=totals['complete_total'])
    ws.cell(row=summary_row + 3, column=1, value="今日未领")
    ws.cell(row=summary_row + 3, column=2, value=totals['pending_total'])
    ws.cell(row=summary_row + 4, column=1, value="TOTAL")
    ws.cell(row=summary_row + 4, column=2, value=f"=SUM(B{summary_row + 2}:B{summary_row + 3})")
    ws.cell(row=summary_row + 5, column=1, value="场口已领")
    ws.cell(row=summary_row + 6, column=1, value="")
    ws.cell(row=summary_row + 7, column=1, value="自动计算")
    #color row C5D9F1 DCE6F1
    ws.cell(row=summary_row + 2, column=1,).fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
    ws.cell(row=summary_row + 3, column=1,).fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
    ws.cell(row=summary_row + 4, column=1,).fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
    ws.cell(row=summary_row + 5, column=1,).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.cell(row=summary_row + 2, column=2,).fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    ws.cell(row=summary_row + 3, column=2,).fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    ws.cell(row=summary_row +4, column=2,).fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    ws.cell(row=summary_row + 5, column=2,).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    #border
    ws.cell(row=summary_row, column=1, ).border  = thin_border
    ws.cell(row=summary_row, column=2, ).border  = thin_border
    ws.cell(row=summary_row + 2, column=1, ).border  = thin_border
    ws.cell(row=summary_row + 2, column=2,).border  = thin_border
    ws.cell(row=summary_row + 3, column=1, ).border  = thin_border
    ws.cell(row=summary_row + 3, column=2, ).border  = thin_border
    ws.cell(row=summary_row + 4, column=1,).border  = thin_border
    ws.cell(row=summary_row + 4, column=2, ).border  = thin_border
    ws.cell(row=summary_row + 5, column=1, ).border  = thin_border
    ws.cell(row=summary_row + 5, column=2, ).border  = thin_border
    ws.cell(row=summary_row + 7, column=1, ).border  = thin_border
    ws.cell(row=summary_row + 7, column=2, ).border  = thin_border

def _add_branch_data(ws, last_row_Bank,date_range):
    """Add branch data to the report"""
    light_pink = "FFD9D9"
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    branches = get_branches(date_range)
    if not branches:
        raise ValueError("No branches found")

    branch_start_row = 4 + last_row_Bank + 8
    lastBranchRow = 0
    for idx, branch in enumerate(branches, start=branch_start_row):
        try:
            # Handle potential None values
            code = getattr(branch, 'code', '') or ''
            total = getattr(branch, 'total', '') or ''
            color_branch = get_color_by_code(code)
            # Branch code
            cell = ws.cell(row=idx, column=1, value=f"{code}".strip())
            cell.fill = PatternFill(start_color=color_branch, end_color=color_branch, fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            
            # Total amount
            cell = ws.cell(row=idx, column=2, value=total)
            cell.fill = PatternFill(start_color=color_branch, end_color=color_branch, fill_type="solid")
            cell.border = thin_border
            lastBranchRow = idx
        except Exception as e:
            logger.error(f"Error adding branch to report: {str(e)}")
            continue
    cell_withdraw = get_column_letter( 2)
    summary_row = last_row_Bank +9
    logger.info(f"Successfully lastBranchRow sii: {last_row_Bank}")
    ws.cell(row=summary_row   , column=2, value=f"=SUM({cell_withdraw}{branch_start_row}:{cell_withdraw}{lastBranchRow+8})")
def get_color_by_code(code):
    """Get hex color code by color reference code"""
    return color_code_mapping.get(code, "Color code not found")